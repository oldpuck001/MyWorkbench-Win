# subtotals.py

import os
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def subtotals_import(request):

    file_path = request.get("data", {}).get("filePath", "")

    sheet_file = pd.ExcelFile(file_path)                            # 使用pandas讀取Excel文件
    sheetnames = sheet_file.sheet_names                             # 獲取所有工作表名稱

    return ['subtotals_import', sheetnames]

def subtotals_index(request):

    file_path = request.get("data", {}).get("filePath", "")
    sheet_name = request.get("data", {}).get("sheetName", "")

    file_extension = os.path.splitext(file_path)[1].lower()

    if file_extension == '.xlsx':
        df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
    elif file_extension == '.xls':
        df = pd.read_excel(file_path, sheet_name=sheet_name, engine='xlrd')

    columns = df.columns.tolist()

    return ['subtotals_index', columns]

def subtotals_generate(request):

    file_path = request.get("data", {}).get("filePath", "")
    save_path = request.get("data", {}).get("savePath", "")
    sheet_name = request.get("data", {}).get("sheet_value", "")
    columnsvalue = request.get("data", {}).get("row_value", "")
    rowvalue = request.get("data", {}).get("column_value", "")
    numbervalue = request.get("data", {}).get("total_value", "")

    if columnsvalue == rowvalue or columnsvalue == numbervalue or rowvalue == numbervalue:

        result_text = {'result_message': '行标题分类列、列项目分类列、合计数值列必须为不通列，请重新选择！'}

        return ['subtotals_generate', result_text]
    
    else:

        file_extension = os.path.splitext(file_path)[1].lower()

        if file_extension == '.xlsx':
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
        elif file_extension == '.xls':
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine='xlrd')

        # 將指定列的空白行轉換為0
        df[numbervalue] = df[numbervalue].replace('', '0')

        # 去除千分位符
        df[numbervalue] = df[numbervalue].astype(str).str.replace(',', '')

        # 將字符格式的數字轉換為數值
        df.loc[0:, numbervalue] = pd.to_numeric(df.loc[0:, numbervalue], errors='coerce').fillna(0)

        exportdf = pd.pivot_table(df.iloc[0:], values=numbervalue, index=rowvalue, columns=columnsvalue, aggfunc='sum')
        exportdf = exportdf.replace('', '0')

        # 创建一个Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active

        # 重設索引並將DataFrame的數據添加到Excel工作表
        exportdf.reset_index(inplace=True)

        # 修改DataFrame的第1列第1行（列标题）为空白
        exportdf.columns.values[0] = ''
        for r in openpyxl.utils.dataframe.dataframe_to_rows(exportdf, index=False, header=True):
            ws.append(r)

        # 调整列宽
        for column_cells in ws.columns:
            length = max(len(str(cell.value))+8 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length

        # 添加合计公式到最后一列
        colnumber = ws.max_column + 1
        row_sum_title = ws.cell(row=1, column=colnumber)
        row_sum_title.value = 'Total'
        for row_index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column)):
            row_sum_cell = ws.cell(row=row_index + 2, column=colnumber)
            row_sum_cell.value = f'=SUM({row[1].coordinate}:{row[-2].coordinate})'
            row_sum_cell.number_format = '#,##0.00'

        # 添加合计公式到最后一行
        rownumber = ws.max_row+1
        col_sum_title = ws.cell(row=rownumber, column=1)
        col_sum_title.value = 'Total'
        for col_index in range(2, ws.max_column + 1):
            col_letter = get_column_letter(col_index)
            col_sum_cell = ws.cell(row=rownumber, column=col_index)
            col_sum_cell.value = f'=SUM({col_letter}2:{col_letter}{ws.max_row-1})'
            col_sum_cell.number_format = '#,##0.00'
            
        # 设置数值单元格格式为两位小数且有千分位符
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1, min_col=2, max_col=ws.max_column):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'  # 设置为带千分位和两位小数

        # 保存Excel文件
        wb.save(save_path)

    result_text = {'result_message': '生成成功！'}

    return ['subtotals_generate', result_text]