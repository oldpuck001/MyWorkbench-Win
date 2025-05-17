# regex.py

import os
import pandas as pd
import openpyxl

def select_file(request):

    file_path = request.get("data", {}).get("filePath", "")

    excelfile = pd.ExcelFile(file_path)                     # 使用pandas讀取Excel文件
    sheetnames = excelfile.sheet_names                      # 獲取所有工作表名稱

    return ['regex_import', sheetnames]


def road_sheet(request):

    file_path = request.get("data", {}).get("filePath", "")
    sheet_name = request.get("data", {}).get("sheetName", "")

    file_extension = os.path.splitext(file_path)[1].lower()

    if file_extension == '.xlsx':
        df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
    elif file_extension == '.xls':
        df = pd.read_excel(file_path, sheet_name=sheet_name, engine='xlrd')

    columns = df.columns.tolist()                           # 获取工作表的列名

    return ['regex_index', columns]


def regex_generate(request):

    sheet_name= request.get("data", {}).get("sheet_value", "")
    column_index = request.get("data", {}).get("column_value", "")
    regex_pattern = request.get("data", {}).get("regex_pattern", "")
    file_path = request.get("data", {}).get("filePath", "")
    save_path = request.get("data", {}).get("savePath", "")
    
    file_extension = os.path.splitext(file_path)[1].lower()

    if file_extension == '.xlsx':
        df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
    elif file_extension == '.xls':
        df = pd.read_excel(file_path, sheet_name=sheet_name, engine='xlrd')

    df.reset_index(drop=True, inplace=True)

    # 使用正則表達式篩選列並更新 DataFrame
    df_result = df[df[column_index].str.contains(regex_pattern, regex=True, na=False)]

    # 创建一个Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active

    # 将DataFrame的数据添加到Excel工作表
    for r in openpyxl.utils.dataframe.dataframe_to_rows(df_result, index=False, header=True):
        ws.append(r)
    
    # 调整列宽
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) + 2 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length

    # 保存Excel文件
    wb.save(save_path)

    return ['regex_generate']