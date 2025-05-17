# import_account_balance_sheet.py

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter

def import_account_balance_sheet(request):

    file_path = request.get("data", {}).get("file_path", "")

    sheet_file = pd.ExcelFile(file_path)                                   # 使用pandas讀取Excel文件
    sheetnames = sheet_file.sheet_names                                    # 獲取所有工作表名稱

    return ['import_account_balance_sheet', sheetnames]


def index_account_balance_sheet(request):

    file_path = request.get("data", {}).get("file_path", "")
    sheet_name = request.get("data", {}).get("sheetName", "")

    file_extension = os.path.splitext(file_path)[1].lower()

    if file_extension == '.xlsx':
        df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
    elif file_extension == '.xls':
        df = pd.read_excel(file_path, sheet_name=sheet_name, engine='xlrd')

    columns = df.columns.tolist()                           # 获取工作表的列名

    return ['index_account_balance_sheet', columns]


def export_account_balance_sheet(request):

    folder_path = request.get('data', {}).get('project_folder', '')
    file_path = request.get('data', {}).get('file_path', '')
    sheet_name = request.get('data', {}).get('sheet_name', '')
    account_id = request.get('data', {}).get('account_id', '')
    account_name = request.get('data', {}).get('account_name', '')
    begin_debit = request.get('data', {}).get('begin_debit', '')
    begin_credit = request.get('data', {}).get('begin_credit', '')
    this_debit = request.get('data', {}).get('this_debit', '')
    this_credit = request.get('data', {}).get('this_credit', '')
    end_debit = request.get('data', {}).get('end_debit', '')
    end_credit = request.get('data', {}).get('end_credit', '')

    file_extension = os.path.splitext(file_path)[1].lower()

    if file_extension == '.xlsx':
        import_account_balance_df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
    elif file_extension == '.xls':
        import_account_balance_df = pd.read_excel(file_path, sheet_name=sheet_name, engine='xlrd')
    else:
        return

    # 将'科目编码'列转换为字符串数据类型
    import_account_balance_df[account_id] = import_account_balance_df[account_id].astype(str)

    # 数据清洗：将数值列的字符串数据类型转换为浮点数类型
    numeric_columns = [begin_debit, begin_credit, this_debit, this_credit, end_debit, end_credit]
    for col in numeric_columns:
        if col in import_account_balance_df.columns:
            import_account_balance_df[col] = import_account_balance_df[col].apply(convert_to_numeric)

    # 获取一级科目代码位数
    min_length = import_account_balance_df[account_id].apply(determine_subject_length).min()

    # 提取后数据（字典）
    export_account_balance_dict = {'一级科目编号': [],
                                   '一级科目名称': [],
                                   '末级科目编号': [],
                                   '末级科目名称': [],
                                   '期初借方': [],
                                   '期初贷方': [],
                                   '本期借方': [],
                                   '本期贷方': [],
                                   '期末借方': [],
                                   '期末贷方': []}

    # 储存已处理的一级科目
    processed_codes = set()

    # 遍历DataFrame提取科目余额表数据
    for index, row in import_account_balance_df.iterrows():
        subject_code = str(row[account_id]).strip()
        if subject_code in processed_codes:
            continue
    
        subject_length = determine_subject_length(subject_code)
        subject_name = row[account_name]
        
        if subject_length == min_length:
            # 将一级科目添加进已处理一级科目集合
            processed_codes.add(subject_code)

            # 添加一级科目行
            export_account_balance_dict['一级科目编号'].append(subject_code)
            export_account_balance_dict['一级科目名称'].append(subject_name)
            export_account_balance_dict['末级科目编号'].append(None)
            export_account_balance_dict['末级科目名称'].append(None)
            export_account_balance_dict['期初借方'].append(row.get(begin_debit, None))
            export_account_balance_dict['期初贷方'].append(row.get(begin_credit, None))
            export_account_balance_dict['本期借方'].append(row.get(this_debit, None))
            export_account_balance_dict['本期贷方'].append(row.get(this_credit, None))
            export_account_balance_dict['期末借方'].append(row.get(end_debit, None))
            export_account_balance_dict['期末贷方'].append(row.get(end_credit, None))

            # 查找所有具有相同一级科目代码的匹配项
            matching_codes = import_account_balance_df[import_account_balance_df[account_id].str.startswith(subject_code)]
            # 求出一级科目对应的末级科目的位数
            max_length = matching_codes[account_id].apply(determine_subject_length).max()
            # 筛选出末级科目数据
            final_level_subjects = matching_codes[matching_codes[account_id].apply(determine_subject_length) == max_length]
            # 如果没有末级科目则进入下一个一级科目
            if matching_codes.empty:
                continue
            # 添加末级科目
            for _, child_row in final_level_subjects.iterrows():
                if str(child_row[account_id]).strip() == subject_code:
                    continue  # 如果遍历到一级科目行则跳过
                    
                export_account_balance_dict['一级科目编号'].append(subject_code)
                export_account_balance_dict['一级科目名称'].append(subject_name)
                export_account_balance_dict['末级科目编号'].append(child_row[account_id])
                export_account_balance_dict['末级科目名称'].append(child_row[account_name])
                export_account_balance_dict['期初借方'].append(child_row.get(begin_debit, None))
                export_account_balance_dict['期初贷方'].append(child_row.get(begin_credit, None))
                export_account_balance_dict['本期借方'].append(child_row.get(this_debit, None))
                export_account_balance_dict['本期贷方'].append(child_row.get(this_credit, None))
                export_account_balance_dict['期末借方'].append(child_row.get(end_debit, None))
                export_account_balance_dict['期末贷方'].append(child_row.get(end_credit, None))

    export_account_balance_df = pd.DataFrame(export_account_balance_dict)

    # 确定保存路径
    export_path = os.path.join(folder_path, '项目数据', '科目余额表.xlsx')

    # 保存为xlsx文件
    export_account_balance_df.to_excel(export_path, index=False, sheet_name='Sheet1', engine='openpyxl')

    # 加载工作簿
    wb = load_workbook(export_path)
    ws = wb.active                              # 默认激活的工作表是第一个，也就是 'Sheet1'

    # 添加边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # 添加筛选功能，筛选行设置在第一行
    ws.auto_filter.ref = ws.dimensions

    # 设置日期和数值格式，遍历所有列
    for col in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in col:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'             # 数字格式设置为千分位和保留两位小数

    # 调整列宽（按实际列顺序）
    column_widths = {
        '一级科目编号': 15,
        '一级科目名称': 20,
        '末级科目编号': 15,
        '末级科目名称': 50,
        '期初借方': 17,
        '期初贷方': 17,
        '本期借方': 17,
        '本期贷方': 17,
        '期末借方': 17,
        '期末贷方': 17
    }

    # 设置列宽
    for i, (col_name, width) in enumerate(column_widths.items(), start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # 保存更改
    wb.save(export_path)
    wb.close()

    return ['export_account_balance_sheet']


def convert_to_numeric(value):
    try:
        return pd.to_numeric(value.replace(',', ''))
    except AttributeError:
        return 0 if pd.isna(value) else value


def determine_subject_length(code):
    return len(str(code).strip())