# import_chronological_account.py

import os
import pandas as pd
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter

def import_chronological_account(request):

    file_path = request.get("data", {}).get("file_path", "")

    sheet_file = pd.ExcelFile(file_path)                                   # 使用pandas讀取Excel文件
    sheetnames = sheet_file.sheet_names                                    # 獲取所有工作表名稱

    return ['import_chronological_account', sheetnames]


def index_chronological_account(request):

    file_path = request.get("data", {}).get("file_path", "")
    sheet_name = request.get("data", {}).get("sheetName", "")

    file_extension = os.path.splitext(file_path)[1].lower()

    if file_extension == '.xlsx':
        df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
    elif file_extension == '.xls':
        df = pd.read_excel(file_path, sheet_name=sheet_name, engine='xlrd')

    columns = df.columns.tolist()                           # 获取工作表的列名

    return ['index_chronological_account', columns]


def export_chronological_account(request):

    folder_path = request.get('data', {}).get('project_folder', '')
    file_path = request.get('data', {}).get('file_path', '')
    sheet_name = request.get('data', {}).get('sheet_name', '')
    account_date = request.get('data', {}).get('account_date', '')
    account_number = request.get('data', {}).get('account_number', '')
    account_name = request.get('data', {}).get('account_name', '')
    account_summary = request.get('data', {}).get('account_summary', '')
    account_debit = request.get('data', {}).get('account_debit', '')
    account_credit = request.get('data', {}).get('account_credit', '')    

    file_extension = os.path.splitext(file_path)[1].lower()

    if file_extension == '.xlsx':
        import_chronological_account_df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
    elif file_extension == '.xls':
        import_chronological_account_df = pd.read_excel(file_path, sheet_name=sheet_name, engine='xlrd')
    else:
        return

    # 对借方、贷方发生额进行数据清洗
    import_chronological_account_df[account_debit] = import_chronological_account_df[account_debit].apply(convert_to_numeric)
    import_chronological_account_df[account_credit] = import_chronological_account_df[account_credit].apply(convert_to_numeric)

    # 转换制单日期为datetime类型并提取月份
    import_chronological_account_df[account_date] = pd.to_datetime(import_chronological_account_df[account_date], errors='coerce')
    import_chronological_account_df['月份'] = import_chronological_account_df[account_date].dt.month

    # 遍历每个凭证号和月份，找出对应的科目
    voucher_groups = import_chronological_account_df.groupby(['月份', account_number])
    for (month, voucher_no), group in voucher_groups:
        primary_subjects = set()
        for subject in group[account_name]:
            primary_subjects.add(str(subject))
        primary_subjects_str = ', '.join(primary_subjects)
        import_chronological_account_df.loc[(import_chronological_account_df['月份'] == month) & (import_chronological_account_df[account_number] == voucher_no), '涉及科目'] = primary_subjects_str

    export_chronological_account_df = pd.DataFrame()
    export_chronological_account_df['涉及科目'] = import_chronological_account_df['涉及科目']
    export_chronological_account_df['日期'] = import_chronological_account_df[account_date]
    export_chronological_account_df['凭证字号'] = import_chronological_account_df[account_number]
    export_chronological_account_df['科目名称'] = import_chronological_account_df[account_name]
    export_chronological_account_df['摘要'] = import_chronological_account_df[account_summary]
    export_chronological_account_df['借方金额'] = import_chronological_account_df[account_debit]
    export_chronological_account_df['贷方金额'] = import_chronological_account_df[account_credit]

    # 确定保存路径
    export_path = os.path.join(folder_path, '项目数据', '序时账.xlsx')

    # 保存为xlsx文件
    export_chronological_account_df.to_excel(export_path, index=False, sheet_name='Sheet1', engine='openpyxl')

    # 加载工作簿
    wb = load_workbook(export_path)
    ws = wb.active                              # 默认激活的工作表是第一个，也就是 'Sheet1'

    # 添加边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # 添加筛选功能，筛选行设置在第一行
    ws.auto_filter.ref = ws.dimensions

    # 设置日期和数值格式，遍历所有列
    for col in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in col:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'             # 数字格式设置为千分位和保留两位小数
            elif isinstance(cell.value, (pd.Timestamp, datetime)):
                cell.value = cell.value.date()              # 强制转换为 date 对象
                cell.number_format = 'yyyy-mm-dd'
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif isinstance(cell.value, date):              # 处理已经转换过的 date 对象
                cell.number_format = 'yyyy-mm-dd'
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # 设置凭证字号列居中
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center', vertical='center')

    # 调整列宽（按实际列顺序）
    column_widths = {
        '涉及科目': 11,
        '日期': 14,
        '凭证字号': 13,
        '科目名称': 50,
        '摘要': 50,
        '借方金额': 15,
        '贷方金额': 15
    }

    # 设置列宽
    for i, (col_name, width) in enumerate(column_widths.items(), start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # 保存更改
    wb.save(export_path)
    wb.close()

    return ['export_chronological_account']


def convert_to_numeric(value):
    try:
        return pd.to_numeric(value.replace(',', ''))
    except AttributeError:
        return 0 if pd.isna(value) else value


def determine_subject_length(code):
    return len(str(code).strip())