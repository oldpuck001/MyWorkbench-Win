# splice_py.py

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

def input_sheet(request):

    file_path = request.get("data", {}).get("filePath", "")

    sheet_file = pd.ExcelFile(file_path)                                   # 使用pandas讀取Excel文件
    sheetnames = sheet_file.sheet_names                                    # 獲取所有工作表名稱

    sheetnames.insert(0, "All worksheets")

    num = request.get("num", "")

    return ['input', sheetnames, num]

def output_sheet(request):

    pathList = [None] * 12
    sheetList = [None] * 12
    dfList = []

    for n in range(12):
        pathList[n] = request.get("data", {}).get(f"file{n}", "")
        sheetList[n] = request.get("data", {}).get(f"sheet{n}", "")

    for n in range(12):
        if pathList[n]:
            if sheetList[n] == 'All worksheets':
                excelfile = pd.ExcelFile(pathList[n])
                sheetnames = excelfile.sheet_names                                    # 獲取所有工作表名稱
                for sheet in sheetnames:
                    dfList.append(pd.read_excel(pathList[n], sheet_name=sheet))
            else:
                dfList.append(pd.read_excel(pathList[n], sheet_name=sheetList[n]))

    df = pd.concat(dfList, axis=0)

    # 创建一个Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active

    # 将DataFrame的数据添加到Excel工作表
    for r in openpyxl.utils.dataframe.dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # 调整列宽
    for column_cells in ws.columns:
        length = max(len(str(cell.value))+8 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length
    
    # 保存Excel文件
    file_path = request.get("data", {}).get("savePath", "")
    wb.save(file_path)

    return ['output']