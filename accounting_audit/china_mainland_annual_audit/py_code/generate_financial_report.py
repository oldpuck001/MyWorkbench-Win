# generate_financial_report.py

import os
import pandas as pd
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate(software_folder):

    basic_info_path = os.path.join(software_folder, 'data_floder', 'basic_info.csv')
    balance_csv_path = os.path.join(software_folder, 'data_floder', 'balance_sheet.csv')
    # income_csv_path = os.path.join(software_folder, 'data_floder', 'income_statement.csv')
    # 现金流量表占位
    # 所有者权益变动表占位
    basic_info_df = pd.read_csv(basic_info_path)
    balance_sheet_df = pd.read_csv(balance_csv_path)
    # income_statement = pd.read_csv(income_csv_path)
    # 现金流量表占位
    # 所有者权益变动表占位

    name = basic_info_df['企业名称'][0]
    # period = basic_info_df['审计期间'][0]
    deadline = basic_info_df['审计截止日'][0]

    # 创建一个新的工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = '资产负债表'

    # 设置标题
    ws.merge_cells('A1:F1')
    ws['A1'].value = '资产负债表'
    ws['A1'].font = Font(size=24, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # 设置报表信息
    ws.merge_cells('A2:B2')
    ws['A2'].value = f'编制单位：{name}'
    ws['A2'].font = Font(size=12)
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells('C2:D2')
    ws['C2'].value = f'          {deadline}'
    ws['C2'].font = Font(size=12)
    ws['C2'].alignment = Alignment(horizontal='left', vertical='center')

    ws['F2'].value = f'单位：元'
    ws['F2'].font = Font(size=12)
    ws['F2'].alignment = Alignment(horizontal='left', vertical='center')

    # 设置列标题
    ws['A3'].value = '项目'
    ws['B3'].value = '期末余额'
    ws['C3'].value = '上年末余额'
    ws['D3'].value = '项目'
    ws['E3'].value = '期末余额'
    ws['F3'].value = '上年末余额'
    for row in ws['A3:F3']:
        for cell in row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    # 报表项目
    balance_project_list = [['流动资产：', '流动负债：'],\
                            ['货币资金', '短期借款'],\
                            ['结算备付金*', '向中央银行借款*'],\
                            ['拆出资金*', '拆入资金*'],\
                            ['交易性金融资产', '交易性金融负债'],\
                            ['△以公允价值计量且其变动计入当期损益的金融资产', '△以公允价值计量且其变动计入当期损益的金融负债'],\
                            ['衍生金融资产', '衍生金融负债'],\
                            ['应收票据', '应付票据'],\
                            ['应收账款', '应付账款'],\
                            ['应收款项融资', '预收款项'],\
                            ['预付款项', '合同负债'],\
                            ['应收保费*', '卖出回购金融资产款*'],\
                            ['应收分保账款*', '吸收存款及同业存放*'],\
                            ['应收分保合同准备金*', '代理买卖证券款*'],\
                            ['其他应收款', '代理承销证券款*'],\
                            ['买入返售金融资产*', '应付职工薪酬'],\
                            ['存货', '应交税费'],\
                            ['合同资产', '其他应付款'],\
                            ['持有待售资产', '应付手续费及佣金*'],\
                            ['一年内到期的非流动资产', '应付分保账款*'],\
                            ['其他流动资产', '持有待售负债'],\
                            ['流动资产合计', '一年内到期的非流动负债'],\
                            ['非流动资产：', '其他流动负债'],\
                            ['发放贷款和垫款*', '流动负债合计'],\
                            ['债权投资', '非流动负债：'],\
                            ['△可供出售金融资产', '保险合同准备金*'],\
                            ['其他债权投资', '长期借款'],\
                            ['△持有至到期投资', '应付债券'],\
                            ['长期应收款', '应付债券：优先股'],\
                            ['长期股权投资', '应付债券：永续债'],\
                            ['其他权益工具投资', '租赁负债'],\
                            ['其他非流动金融资产', '长期应付款'],\
                            ['投资性房地产', '预计负债'],\
                            ['固定资产', '递延收益'],\
                            ['在建工程', '递延所得税负债'],\
                            ['生产性生物资产', '其他非流动负债'],\
                            ['油气资产', '非流动负债合计'],\
                            ['使用权资产', '负债合计'],\
                            ['无形资产', '所有者权益（或股东权益）：'],\
                            ['开发支出', '实收资本（或股本）'],\
                            ['商誉', '#减：已归还投资'],\
                            ['长期待摊费用', '#实收资本（或股本）净额'],\
                            ['递延所得税资产', '其他权益工具'],\
                            ['其他非流动资产', '其他权益工具：优先股'],\
                            ['非流动资产合计', '其他权益工具：永续债'],\
                            ['', '资本公积'],\
                            ['', '减：库存股'],\
                            ['', '其他综合收益'],\
                            ['', '专项储备'],\
                            ['', '盈余公积'],\
                            ['', '一般风险准备*'],\
                            ['', '未分配利润'],\
                            ['', '*归属于母公司所有者权益（或股东权益）合计'],\
                            ['', '*少数股东权益'],\
                            ['', '所有者权益（或股东权益）合计'],\
                            ['资产总计', '负债和所有者权益（或股东权益）总计']]

    start_row = 4
    for p in balance_project_list:
        left = p[0]
        if left == '':
            pass
        elif left[-1] in ['：']:
            ws.cell(row=start_row, column=1, value=left)
        elif left in ['流动资产合计','非流动资产合计', '资产总计', '流动负债合计', '非流动负债合计', '负债合计', '#实收资本（或股本）净额', '*归属于母公司所有者权益（或股东权益）合计', '所有者权益（或股东权益）合计', '负债和所有者权益（或股东权益）总计']:
            ws.cell(row=start_row, column=1, value=left)
            if left == '流动资产合计':
                left_end = f"=SUM({ws.cell(row=start_row - 20, column=ws.cell(row=start_row, column=2).column).coordinate}:{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=2).column).coordinate})"
                left_previous = f"=SUM({ws.cell(row=start_row - 20, column=ws.cell(row=start_row, column=3).column).coordinate}:{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=3).column).coordinate})"
            elif left == '非流动资产合计':
                left_end = f"=SUM({ws.cell(row=start_row - 21, column=ws.cell(row=start_row, column=2).column).coordinate}:{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=2).column).coordinate})"
                left_previous = f"=SUM({ws.cell(row=start_row - 21, column=ws.cell(row=start_row, column=3).column).coordinate}:{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=3).column).coordinate})"
            elif left == '资产总计':
                left_end = f"={ws.cell(row=start_row - 34, column=ws.cell(row=start_row, column=2).column).coordinate}+{ws.cell(row=start_row - 11, column=ws.cell(row=start_row, column=2).column).coordinate}"
                left_previous = f"={ws.cell(row=start_row - 34, column=ws.cell(row=start_row, column=3).column).coordinate}+{ws.cell(row=start_row - 11, column=ws.cell(row=start_row, column=3).column).coordinate}"
            elif left == '流动负债合计':
                left_end = f"=SUM({ws.cell(row=start_row - 22, column=ws.cell(row=start_row, column=2).column).coordinate}:{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=2).column).coordinate})"
                left_previous = f"=SUM({ws.cell(row=start_row - 22, column=ws.cell(row=start_row, column=3).column).coordinate}:{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=3).column).coordinate})"
            elif left == '非流动负债合计':
                left_end = f"=SUM({ws.cell(row=start_row - 11, column=ws.cell(row=start_row, column=2).column).coordinate}:{ws.cell(row=start_row - 9, column=ws.cell(row=start_row, column=2).column).coordinate})+SUM({ws.cell(row=start_row - 6, column=ws.cell(row=start_row, column=2).column).coordinate}:{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=2).column).coordinate})"
                left_previous = f"=SUM({ws.cell(row=start_row - 11, column=ws.cell(row=start_row, column=3).column).coordinate}:{ws.cell(row=start_row - 9, column=ws.cell(row=start_row, column=3).column).coordinate})+SUM({ws.cell(row=start_row - 6, column=ws.cell(row=start_row, column=3).column).coordinate}:{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=3).column).coordinate})"
            elif left == '负债合计':
                left_end = f"={ws.cell(row=start_row - 14, column=ws.cell(row=start_row, column=2).column).coordinate}+{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=2).column).coordinate}"
                left_previous = f"={ws.cell(row=start_row - 14, column=ws.cell(row=start_row, column=3).column).coordinate}+{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=3).column).coordinate}"
            elif left == '#实收资本（或股本）净额':
                left_end = f"={ws.cell(row=start_row - 2, column=ws.cell(row=start_row, column=2).column).coordinate}-{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=2).column).coordinate}"
                left_previous = f"={ws.cell(row=start_row - 2, column=ws.cell(row=start_row, column=3).column).coordinate}-{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=3).column).coordinate}"
            elif left == '*归属于母公司所有者权益（或股东权益）合计':
                left_end = f"=SUM({ws.cell(row=start_row - 11, column=ws.cell(row=start_row, column=2).column).coordinate}:{ws.cell(row=start_row - 10, column=ws.cell(row=start_row, column=2).column).coordinate})+SUM({ws.cell(row=start_row - 7, column=ws.cell(row=start_row, column=2).column).coordinate}:{ws.cell(row=start_row - 7, column=ws.cell(row=start_row, column=2).column).coordinate})-SUM({ws.cell(row=start_row - 6, column=ws.cell(row=start_row, column=2).column).coordinate}:{ws.cell(row=start_row - 6, column=ws.cell(row=start_row, column=2).column).coordinate})+SUM({ws.cell(row=start_row - 5, column=ws.cell(row=start_row, column=2).column).coordinate}:{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=2).column).coordinate})"
                left_previous = f"=SUM({ws.cell(row=start_row - 11, column=ws.cell(row=start_row, column=3).column).coordinate}:{ws.cell(row=start_row - 10, column=ws.cell(row=start_row, column=3).column).coordinate})+SUM({ws.cell(row=start_row - 7, column=ws.cell(row=start_row, column=3).column).coordinate}:{ws.cell(row=start_row - 7, column=ws.cell(row=start_row, column=3).column).coordinate})-SUM({ws.cell(row=start_row - 6, column=ws.cell(row=start_row, column=3).column).coordinate}:{ws.cell(row=start_row - 6, column=ws.cell(row=start_row, column=3).column).coordinate})+SUM({ws.cell(row=start_row - 5, column=ws.cell(row=start_row, column=3).column).coordinate}:{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=3).column).coordinate})"
            elif left == '所有者权益（或股东权益）合计':
                left_end = f"=SUM({ws.cell(row=start_row - 2, column=ws.cell(row=start_row, column=2).column).coordinate}:{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=2).column).coordinate})"
                left_previous = f"=SUM({ws.cell(row=start_row - 2, column=ws.cell(row=start_row, column=3).column).coordinate}:{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=3).column).coordinate})"
            elif left == '负债和所有者权益（或股东权益）总计':
                left_end = f"={ws.cell(row=start_row - 18, column=ws.cell(row=start_row, column=2).column).coordinate}+{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=2).column).coordinate}"
                left_previous = f"={ws.cell(row=start_row - 18, column=ws.cell(row=start_row, column=3).column).coordinate}+{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=3).column).coordinate}"
            ws.cell(row=start_row, column=2, value=left_end)
            ws.cell(row=start_row, column=3, value=left_previous)
        else:
            if left == '应付债券：优先股' or left == '其他权益工具：优先股':
                ws.cell(row=start_row, column=1, value='  其中：优先股')
            elif left == '应付债券：永续债' or left == '其他权益工具：永续债':
                ws.cell(row=start_row, column=1, value='        永续债')
            else:
                ws.cell(row=start_row, column=1, value=left)
            left_end = balance_sheet_df[left][8] if balance_sheet_df[left][8] != 0 else ''
            ws.cell(row=start_row, column=2, value=left_end)
            left_previous = balance_sheet_df[left][3] if balance_sheet_df[left][3] != 0 else ''
            ws.cell(row=start_row, column=3, value=left_previous)
        right = p[1]
        if right == '':
            pass
        elif right[-1] in ['：']:
            ws.cell(row=start_row, column=4, value=right)
        elif right in ['流动资产合计','非流动资产合计', '资产总计', '流动负债合计', '非流动负债合计', '负债合计', '#实收资本（或股本）净额', '*归属于母公司所有者权益（或股东权益）合计', '所有者权益（或股东权益）合计', '负债和所有者权益（或股东权益）总计']:
            ws.cell(row=start_row, column=4, value=right)
            if right == '流动资产合计':
                right_end = f"=SUM({ws.cell(row=start_row - 20, column=ws.cell(row=start_row, column=5).column).coordinate}:{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=5).column).coordinate})"
                right_previous = f"=SUM({ws.cell(row=start_row - 20, column=ws.cell(row=start_row, column=6).column).coordinate}:{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=6).column).coordinate})"
            elif right == '非流动资产合计':
                right_end = f"=SUM({ws.cell(row=start_row - 21, column=ws.cell(row=start_row, column=5).column).coordinate}:{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=5).column).coordinate})"
                right_previous = f"=SUM({ws.cell(row=start_row - 21, column=ws.cell(row=start_row, column=6).column).coordinate}:{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=6).column).coordinate})"
            elif right == '资产总计':
                right_end = f"={ws.cell(row=start_row - 34, column=ws.cell(row=start_row, column=5).column).coordinate}+{ws.cell(row=start_row - 11, column=ws.cell(row=start_row, column=5).column).coordinate}"
                right_previous = f"={ws.cell(row=start_row - 34, column=ws.cell(row=start_row, column=6).column).coordinate}+{ws.cell(row=start_row - 11, column=ws.cell(row=start_row, column=6).column).coordinate}"
            elif right == '流动负债合计':
                right_end = f"=SUM({ws.cell(row=start_row - 22, column=ws.cell(row=start_row, column=5).column).coordinate}:{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=5).column).coordinate})"
                right_previous = f"=SUM({ws.cell(row=start_row - 22, column=ws.cell(row=start_row, column=6).column).coordinate}:{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=6).column).coordinate})"
            elif right == '非流动负债合计':
                right_end = f"=SUM({ws.cell(row=start_row - 11, column=ws.cell(row=start_row, column=5).column).coordinate}:{ws.cell(row=start_row - 9, column=ws.cell(row=start_row, column=5).column).coordinate})+SUM({ws.cell(row=start_row - 6, column=ws.cell(row=start_row, column=5).column).coordinate}:{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=5).column).coordinate})"
                right_previous = f"=SUM({ws.cell(row=start_row - 11, column=ws.cell(row=start_row, column=6).column).coordinate}:{ws.cell(row=start_row - 9, column=ws.cell(row=start_row, column=6).column).coordinate})+SUM({ws.cell(row=start_row - 6, column=ws.cell(row=start_row, column=6).column).coordinate}:{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=6).column).coordinate})"
            elif right == '负债合计':
                right_end = f"={ws.cell(row=start_row - 14, column=ws.cell(row=start_row, column=5).column).coordinate}+{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=5).column).coordinate}"
                right_previous = f"={ws.cell(row=start_row - 14, column=ws.cell(row=start_row, column=6).column).coordinate}+{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=6).column).coordinate}"
            elif right == '#实收资本（或股本）净额':
                right_end = f"={ws.cell(row=start_row - 2, column=ws.cell(row=start_row, column=5).column).coordinate}-{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=5).column).coordinate}"
                right_previous = f"={ws.cell(row=start_row - 2, column=ws.cell(row=start_row, column=6).column).coordinate}-{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=6).column).coordinate}"
            elif right == '*归属于母公司所有者权益（或股东权益）合计':
                right_end = f"=SUM({ws.cell(row=start_row - 11, column=ws.cell(row=start_row, column=5).column).coordinate}:{ws.cell(row=start_row - 10, column=ws.cell(row=start_row, column=5).column).coordinate})+SUM({ws.cell(row=start_row - 7, column=ws.cell(row=start_row, column=5).column).coordinate}:{ws.cell(row=start_row - 7, column=ws.cell(row=start_row, column=5).column).coordinate})-SUM({ws.cell(row=start_row - 6, column=ws.cell(row=start_row, column=5).column).coordinate}:{ws.cell(row=start_row - 6, column=ws.cell(row=start_row, column=5).column).coordinate})+SUM({ws.cell(row=start_row - 5, column=ws.cell(row=start_row, column=5).column).coordinate}:{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=5).column).coordinate})"
                right_previous = f"=SUM({ws.cell(row=start_row - 11, column=ws.cell(row=start_row, column=6).column).coordinate}:{ws.cell(row=start_row - 10, column=ws.cell(row=start_row, column=6).column).coordinate})+SUM({ws.cell(row=start_row - 7, column=ws.cell(row=start_row, column=6).column).coordinate}:{ws.cell(row=start_row - 7, column=ws.cell(row=start_row, column=6).column).coordinate})-SUM({ws.cell(row=start_row - 6, column=ws.cell(row=start_row, column=6).column).coordinate}:{ws.cell(row=start_row - 6, column=ws.cell(row=start_row, column=6).column).coordinate})+SUM({ws.cell(row=start_row - 5, column=ws.cell(row=start_row, column=6).column).coordinate}:{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=6).column).coordinate})"
            elif right == '所有者权益（或股东权益）合计':
                right_end = f"=SUM({ws.cell(row=start_row - 2, column=ws.cell(row=start_row, column=5).column).coordinate}:{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=5).column).coordinate})"
                right_previous = f"=SUM({ws.cell(row=start_row - 2, column=ws.cell(row=start_row, column=6).column).coordinate}:{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=6).column).coordinate})"
            elif right == '负债和所有者权益（或股东权益）总计':
                right_end = f"={ws.cell(row=start_row - 18, column=ws.cell(row=start_row, column=5).column).coordinate}+{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=5).column).coordinate}"
                right_previous = f"={ws.cell(row=start_row - 18, column=ws.cell(row=start_row, column=6).column).coordinate}+{ws.cell(row=start_row - 1, column=ws.cell(row=start_row, column=6).column).coordinate}"
            ws.cell(row=start_row, column=5, value=right_end)
            ws.cell(row=start_row, column=6, value=right_previous)
        else:
            if right == '应付债券：优先股' or right == '其他权益工具：优先股':
                ws.cell(row=start_row, column=4, value='  其中：优先股')
            elif right == '应付债券：永续债' or right == '其他权益工具：永续债':
                ws.cell(row=start_row, column=4, value='        永续债')
            else:
                ws.cell(row=start_row, column=4, value=right)
            right_end =balance_sheet_df[right][8] if balance_sheet_df[right][8] != 0 else ''
            ws.cell(row=start_row, column=5, value=right_end)
            right_previous = balance_sheet_df[right][3] if balance_sheet_df[right][3] != 0 else ''
            ws.cell(row=start_row, column=6, value=right_previous)
        start_row += 1

    # 设置单元格格式为数值，保留两位小数，并显示千分位符
    for row in ws['B4:C59']:
        for cell in row:
            cell.number_format = '#,##0.00'
    for row in ws['E4:F59']:
        for cell in row:
            cell.number_format = '#,##0.00'

    ws['A25'].font = Font(bold=True)
    ws['B25'].font = Font(bold=True)
    ws['C25'].font = Font(bold=True)
    ws['A48'].font = Font(bold=True)
    ws['B48'].font = Font(bold=True)
    ws['C48'].font = Font(bold=True)
    ws['A59'].font = Font(bold=True)
    ws['B59'].font = Font(bold=True)
    ws['C59'].font = Font(bold=True)

    ws['D27'].font = Font(bold=True)
    ws['E27'].font = Font(bold=True)
    ws['F27'].font = Font(bold=True)
    ws['D40'].font = Font(bold=True)
    ws['E40'].font = Font(bold=True)
    ws['F40'].font = Font(bold=True)
    ws['D41'].font = Font(bold=True)
    ws['E41'].font = Font(bold=True)
    ws['F41'].font = Font(bold=True)
    ws['D58'].font = Font(bold=True)
    ws['E58'].font = Font(bold=True)
    ws['F58'].font = Font(bold=True)
    ws['D59'].font = Font(bold=True)
    ws['E59'].font = Font(bold=True)
    ws['F59'].font = Font(bold=True)

    # 设置A9、D9、D59三个单元格的自动换行
    ws['A9'].alignment = Alignment(wrap_text=True)
    ws['D9'].alignment = Alignment(wrap_text=True)
    ws['D56'].alignment = Alignment(wrap_text=True)
    ws['D59'].alignment = Alignment(wrap_text=True)

    # 设置列宽
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18

    # 添加边框样式
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=3, min_col=1, max_row=59, max_col=6):
        for cell in row:
            cell.border = thin_border

    statement_xlsx_path = os.path.join(software_folder, 'report', '会计报表.xlsx')
    wb.save(statement_xlsx_path)

    return statement_xlsx_path