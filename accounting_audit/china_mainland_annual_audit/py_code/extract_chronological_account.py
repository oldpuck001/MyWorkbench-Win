# extract_chronological_account.py

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

def extract_chronological_account_xlsx(filepath, data_folder):
    
    file_extension = os.path.splitext(filepath)[1].lower()
    if file_extension == '.xlsx':
        chronological_account_df = pd.read_excel(filepath, engine='openpyxl')
    elif file_extension == '.xls':
        chronological_account_df = pd.read_excel(filepath, engine='xlrd')

    chronological_account_df['借方'] = chronological_account_df['借方'].apply(convert_to_numeric)
    chronological_account_df['贷方'] = chronological_account_df['贷方'].apply(convert_to_numeric)

    # 转换制单日期为datetime类型并提取月份
    chronological_account_df['日期'] = pd.to_datetime(chronological_account_df['日期'], errors='coerce')
    chronological_account_df['月份'] = chronological_account_df['日期'].dt.month

    # 遍历每个凭证号和月份，找出对应的一级科目
    voucher_groups = chronological_account_df.groupby(['月份', '凭证字号'])
    for (month, voucher_no), group in voucher_groups:
        primary_subjects = set()
        for subject in group['一级科目']:
            primary_subjects.add(str(subject))
        primary_subjects_str = ', '.join(primary_subjects)
        chronological_account_df.loc[(chronological_account_df['月份'] == month) & (chronological_account_df['凭证字号'] == voucher_no), '涉及科目'] = primary_subjects_str

    chronological_account_df = chronological_account_df.drop(columns=['一级科目'])
    chronological_account_df = chronological_account_df.drop(columns=['月份'])

    # 保存处理后的文件
    save_path = os.path.join(data_folder, 'chronological_account_data.xlsx')
    chronological_account_df.to_excel(save_path, index=False)
    
    # 加载工作簿
    wb = load_workbook(save_path)
    ws = wb.active  # 默认激活的工作表是第一个，也就是 'Sheet1'

    # 添加边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # 添加筛选功能，筛选行设置在第一行
    ws.auto_filter.ref = ws.dimensions

    # 设置日期和数值格式，遍历所有列
    for col in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in col:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'  # 数字格式设置为千分位和保留两位小数
            elif isinstance(cell.value, pd.Timestamp):
                cell.number_format = 'YYYY-MM-DD'  # 确保只显示日期部分

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # 获取列字母
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # 保存更改
    wb.save(save_path)
    wb.close()

    return {'message': '导入成功。'}, save_path

def convert_to_numeric(value):
    try:
        return pd.to_numeric(value.replace(',', ''))
    except AttributeError:
        return 0 if pd.isna(value) else value