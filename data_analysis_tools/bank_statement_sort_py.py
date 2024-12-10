# bank_statement_sort_py.py

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side

def bank_statement_sort_import(request):

    file_path = request.get("data", {}).get("filePath", "")

    sheet_file = pd.ExcelFile(file_path)                                   # 使用pandas讀取Excel文件
    sheetnames = sheet_file.sheet_names                                    # 獲取所有工作表名稱

    return ['bank_statement_sort_import', sheetnames]

def bank_statement_sort_index(request):

    file_path = request.get("data", {}).get("filePath", "")
    sheet_name = request.get("data", {}).get("sheetName", "")

    df = pd.read_excel(file_path, sheet_name=sheet_name)    # 读取指定的工作表
    columns = df.columns.tolist()                           # 获取工作表的列名

    return ['bank_statement_sort_index', columns]

def bank_statement_sort_debit_or_credit(request):

    file_path = request.get("data", {}).get("filePath", "")
    sheet_name = request.get("data", {}).get("sheetName", "")
    column_name = request.get("data", {}).get("columnName", "")

    debit_or_credit_list_cleaned =[]

    df = pd.read_excel(file_path, sheet_name=sheet_name)    # 读取指定的工作表
    debit_or_credit_np = df[column_name].unique()
    debit_or_credit_list = debit_or_credit_np.tolist()

    for item in debit_or_credit_list:
        if not isinstance(item, (int, float)):
            debit_or_credit_list_cleaned.append(item.replace('\t', ''))
        else:
            debit_or_credit_list_cleaned.append(item)

    return ['bank_statement_sort_debit_or_credit', debit_or_credit_list_cleaned]

def bank_statement_sort_export(request):
    
    source_file_path = request.get("data", {}).get("filePath", "")
    sheet_name = request.get("data", {}).get("sheet_name", "")
    debit_or_credit_column = request.get("data", {}).get("debit_or_credit_column", "")
    credit_column = request.get("data", {}).get("credit_column", "")
    debit_column = request.get("data", {}).get("debit_column", "")
    name_column = request.get("data", {}).get("name_column", "")
    bank_column = request.get("data", {}).get("bank_column", "")
    number_column = request.get("data", {}).get("number_column", "")
    value_column = request.get("data", {}).get("value_column", "")
    credit_priority = request.get("data", {}).get("credit_priority", "")
    credit_priority_list = [name.strip() for name in credit_priority.split(',')]
    debit_priority = request.get("data", {}).get("debit_priority", "")
    debit_priority_list = [name.strip() for name in debit_priority.split(',')]

    file_path = os.path.dirname(source_file_path)
    file_name = os.path.splitext(source_file_path)[0]
    target_file_path = os.path.join(file_path, f'{file_name}银行对账单分类.xlsx')

    credit_name_summary = {'credit_name_sort': [], 'credit_name_value': []}
    credit_name_number_summary_dict = {}
    credit_name_number_summary = {'credit_name_number_bank': [], 'credit_name_number_sort': [], 'credit_name_number_value': []}
    debit_name_summary = {'debit_name_sort': [], 'debit_name_value': []}
    debit_name_number_summary_dict = {}
    debit_name_number_summary = {'debit_name_number_bank': [], 'debit_name_number_sort': [], 'debit_name_number_value': []}

    if debit_or_credit_column == value_column:
        return ['single_sort_export_no']
    else:
        file_extension = os.path.splitext(source_file_path)[1].lower()
        if file_extension == '.xlsx':
            df = pd.read_excel(source_file_path, sheet_name=sheet_name, engine='openpyxl')
        elif file_extension == '.xls':
            df = pd.read_excel(source_file_path, sheet_name=sheet_name, engine='xlrd')

        # 將指定列的空白行轉換為0
        df[value_column] = df[value_column].replace('', '0')

        # 去除千分位符
        df[value_column] = df[value_column].astype(str).str.replace(',', '')

        # 將字符格式的數字轉換為數值
        df.loc[0:, value_column] = pd.to_numeric(df.loc[0:, value_column], errors='coerce').fillna(0)

        df.loc[:, name_column] = df[name_column].fillna('<空白>').str.strip().replace('', '<空白>')
        df.loc[:, bank_column] = df[bank_column].fillna('<空白>').str.strip().replace('', '<空白>')
        df.loc[:, number_column] = df[number_column].fillna('<空白>').replace('', '<空白>')

        # 第一层分组，划分转入与转出
        df[debit_or_credit_column] = df[debit_or_credit_column].astype(str).str.strip()
        df_grouped = df.groupby(debit_or_credit_column)
        credit_df = df_grouped.get_group(credit_column)
        debit_df = df_grouped.get_group(debit_column)
        credit_summary_total = credit_df[value_column].sum()
        debit_summary_total = debit_df[value_column].sum()

        # 第二层分组，按对方户名进行分组
        # 第二层分组的转入部分
        credit_name_sorts = credit_df[name_column].unique()
        credit_name_dfs = {credit_name_sort: credit_df.loc[credit_df[name_column] == credit_name_sort] for credit_name_sort in credit_name_sorts}

        for credit_name_sort, credit_name_df in credit_name_dfs.items():
            credit_name_sort_cleaned = credit_name_sort.replace('\t', '') if not isinstance(credit_name_sort, float) else credit_name_sort
            credit_name_total = credit_name_df[value_column].sum()
            credit_name_summary['credit_name_sort'].append(credit_name_sort_cleaned)
            credit_name_summary['credit_name_value'].append(credit_name_total)

            # 第三层分组，按账户进行分组
            # 第三层分组的转入部分
            credit_name_number_sorts = credit_name_df[number_column].unique()
            credit_name_number_dfs = {credit_name_number_sort: credit_name_df.loc[credit_name_df[number_column] == credit_name_number_sort] for credit_name_number_sort in credit_name_number_sorts}

            for credit_name_number_sort, credit_name_number_df in credit_name_number_dfs.items():
                credit_bank_name = credit_name_number_df[bank_column].unique()
                credit_bank_name_cleaned = [item.replace('\t', '') for item in credit_bank_name if not isinstance(item, float)]
                credit_number_total = credit_name_number_df[value_column].sum()
                credit_name_number_summary['credit_name_number_bank'].append(credit_bank_name_cleaned)
                credit_name_number_summary['credit_name_number_sort'].append(credit_name_number_sort)
                credit_name_number_summary['credit_name_number_value'].append(credit_number_total)

            credit_name_number_summary_dict[credit_name_sort_cleaned] = pd.DataFrame(credit_name_number_summary).sort_values(by='credit_name_number_value', ascending=False)

            credit_name_number_summary['credit_name_number_bank'].clear()
            credit_name_number_summary['credit_name_number_sort'].clear()
            credit_name_number_summary['credit_name_number_value'].clear()

        credit_name_summary_df_not_sorted = pd.DataFrame(credit_name_summary)
        credit_name_summary_df = credit_name_summary_df_not_sorted.sort_values(by='credit_name_value', ascending=False)

        # 第二层分组的转出部分
        debit_name_sorts = debit_df[name_column].unique()
        debit_name_dfs = {debit_name_sort: debit_df.loc[debit_df[name_column] == debit_name_sort] for debit_name_sort in debit_name_sorts}

        for debit_name_sort, debit_name_df in debit_name_dfs.items():
            debit_name_sort_cleaned = debit_name_sort.replace('\t', '') if not isinstance(debit_name_sort, float) else debit_name_sort
            debit_name_total = debit_name_df[value_column].sum()
            debit_name_summary['debit_name_sort'].append(debit_name_sort_cleaned)
            debit_name_summary['debit_name_value'].append(debit_name_total)

            # 第三层分组，按账户进行分组
            # 第三层分组的转入部分
            debit_name_number_sorts = debit_name_df[number_column].unique()
            debit_name_number_dfs = {debit_name_number_sort: debit_name_df.loc[debit_name_df[number_column] == debit_name_number_sort] for debit_name_number_sort in debit_name_number_sorts}

            for debit_name_number_sort, debit_name_number_df in debit_name_number_dfs.items():
                debit_bank_name = debit_name_number_df[bank_column].unique()
                debit_bank_name_cleaned = [item.replace('\t', '') for item in debit_bank_name if not isinstance(item, float)]
                debit_number_total = debit_name_number_df[value_column].sum()
                debit_name_number_summary['debit_name_number_bank'].append(debit_bank_name_cleaned)
                debit_name_number_summary['debit_name_number_sort'].append(debit_name_number_sort)
                debit_name_number_summary['debit_name_number_value'].append(debit_number_total)

            debit_name_number_summary_dict[debit_name_sort_cleaned] = pd.DataFrame(debit_name_number_summary).sort_values(by='debit_name_number_value', ascending=False)

            debit_name_number_summary['debit_name_number_bank'].clear()
            debit_name_number_summary['debit_name_number_sort'].clear()
            debit_name_number_summary['debit_name_number_value'].clear()

        debit_name_summary_df_not_sorted = pd.DataFrame(debit_name_summary)
        debit_name_summary_df = debit_name_summary_df_not_sorted.sort_values(by='debit_name_value', ascending=False)

        # 创建 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "树状图"

        # 添加贷方数据到左边
        credit_cycle_not_sorted = list(credit_name_summary_df['credit_name_sort'])
        ordered = [item for item in credit_priority_list if item in credit_cycle_not_sorted]
        remaining = [item for item in credit_cycle_not_sorted if item not in ordered]
        credit_cycle = ordered + remaining

        credit_step = 1
        credit_skip = 0

        for credit_name in credit_cycle:
            credit_step += credit_skip
            a_cell = ws.cell(row=credit_step, column=1, value=credit_name)
            cell_format(a_cell)
            credit_name_value = credit_name_summary_df.query('credit_name_sort == @credit_name')['credit_name_value'].values[0]
            b_cell = ws.cell(row=credit_step, column=2, value=credit_name_value)
            cell_format(b_cell)

            credit_skip = len(credit_name_number_summary_dict[credit_name]) + 1

            secondary_credit_step = credit_step
            for row in dataframe_to_rows(credit_name_number_summary_dict[credit_name], index=False, header=False):
                if isinstance(row[0], list):
                    value = ', '.join(str(v) for v in row[0]) if row[0] else '<空白>'
                c_cell = ws.cell(row=secondary_credit_step, column=3, value=value)
                cell_format(c_cell)
                d_cell = ws.cell(row=secondary_credit_step, column=4, value=row[1])
                cell_format(d_cell)
                e_cell = ws.cell(row=secondary_credit_step, column=5, value=row[2])
                cell_format(e_cell)
                secondary_credit_step += 1

        # 中间部分
        f1_cell = ws.cell(row=1, column=6, value='汇入金额合计')
        cell_format(f1_cell)
        f2_cell = ws.cell(row=2, column=6, value=credit_summary_total)
        cell_format(f2_cell)

        g1_cell = ws.cell(row=1, column=7, value='汇出金额合计')
        cell_format(g1_cell)
        g2_cell = ws.cell(row=2, column=7, value=debit_summary_total)
        cell_format(g2_cell)

        # 添加借方数据到右边
        debit_cycle_not_sorted = list(debit_name_summary_df['debit_name_sort'])
        ordered = [item for item in debit_priority_list if item in debit_cycle_not_sorted]
        remaining = [item for item in debit_cycle_not_sorted if item not in ordered]
        debit_cycle = ordered + remaining

        debit_step = 1
        debit_skip = 0

        for debit_name in debit_cycle:
            debit_step += debit_skip
            h_cell = ws.cell(row=debit_step, column=8, value=debit_name)
            cell_format(h_cell)
            debit_name_value = debit_name_summary_df.query('debit_name_sort == @debit_name')['debit_name_value'].values[0]
            i_cell = ws.cell(row=debit_step, column=9, value=debit_name_value)
            cell_format(i_cell)

            debit_skip = len(debit_name_number_summary_dict[debit_name]) + 1
            
            secondary_debit_step = debit_step
            for row in dataframe_to_rows(debit_name_number_summary_dict[debit_name], index=False, header=False):
                if isinstance(row[0], list):
                    value = ', '.join(str(v) for v in row[0]) if row[0] else '<空白>'
                j_cell = ws.cell(row=secondary_debit_step, column=10, value=value)
                cell_format(j_cell)
                k_cell = ws.cell(row=secondary_debit_step, column=11, value=row[1])
                cell_format(k_cell)
                l_cell = ws.cell(row=secondary_debit_step, column=12, value=row[2])
                cell_format(l_cell)
                secondary_debit_step += 1

        # 保存 Excel 文件
        wb.save(target_file_path)

    return ['bank_statement_sort_export']

def cell_format(format_cell):

    # 设置千分位符和两位小数的数字格式
    format_cell.number_format = '#,##0.00'

    # 设置单元格边框（上下左右）
    thin_boeder = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    
    format_cell.border = thin_boeder

    return